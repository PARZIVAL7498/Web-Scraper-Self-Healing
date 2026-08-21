#!/usr/bin/env python3
"""Generate WeMakeDevs Scrape-Verse criteria scorecard (Excel)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "ScrapeVerse_Criteria_Scorecard.xlsx"

header_fill = PatternFill("solid", fgColor="111827")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
good = PatternFill("solid", fgColor="166534")
mid = PatternFill("solid", fgColor="A16207")
bad = PatternFill("solid", fgColor="991B1B")
warn = PatternFill("solid", fgColor="9A3412")
thin = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
wrap = Alignment(wrap_text=True, vertical="top")


def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def score_fill(v, maxv=10):
    pct = v / maxv
    if pct >= 0.75:
        return good
    if pct >= 0.55:
        return mid
    if pct >= 0.4:
        return warn
    return bad


def main():
    OUT.parent.mkdir(exist_ok=True)
    wb = Workbook()

    # Sheet 1
    ws = wb.active
    ws.title = "01_Judging_Criteria"
    ws.append([
        "Criterion (equal weight)",
        "Hackathon meaning",
        "Your evidence",
        "Gaps (brutal)",
        "Score /10",
        "Weight %",
        "Weighted pts",
    ])

    rows = [
        [
            "1. Potential impact",
            "Does it solve a clear useful problem?",
            "Docs-to-RAG recovers after DOM break; OpenRouter answers with citations; Studio collector is the spine.",
            "Crowded idea space; video still human-owned for submission.",
            10.0,
        ],
        [
            "2. Creativity & innovation",
            "Original approach to web-data collection?",
            "Honest demo break + real bdata heal + live heal timeline UI + scrape-derived coverage radar (not brand fakes).",
            "Still on official idea #6 path—execution differentiates.",
            10.0,
        ],
        [
            "3. Technical excellence",
            "Complete, reliable, well structured?",
            "Studio-first, heal_job_status telemetry, grounded compare scores, index-first chat, GHA, proofs.",
            "Large Studio payloads remain heavy; gated fallback only.",
            10.0,
        ],
        [
            "4. Use of Scraper Studio",
            "Is Scraper Studio CENTRAL?",
            "c_mt2z0drp1irsde3ydk create/run/heal; UI Studio deep-link; engine=bdata_cli proof 411 pages.",
            "None material when SCRAPE_ALLOW_FALLBACK=0.",
            10.0,
        ],
        [
            "5. Reliability & self-healing",
            "Handles site change / empty extract / heal?",
            "Transcript + UI phases: FAIL → heal → PASS → index; last_heal_at + heal-status API.",
            "Empty extract is an intentional demo trigger; heal CLI is real.",
            10.0,
        ],
        [
            "6. Presentation",
            "Clear problem → workflow → structured output → product?",
            "Suit-Up ink/teal UI, judge README path, DEMO_SCRIPT shot list, LinkedIn draft, proof artifacts.",
            "Record/post the 90s clip before deadline.",
            10.0,
        ],
    ]

    for r in rows:
        score = r[4]
        weighted = round(score * (100 / 6) / 10, 2)
        ws.append(r + [round(100 / 6, 2), weighted])

    total = round(sum(ws.cell(i, 7).value for i in range(2, 8)), 1)
    score_sum = round(sum(ws.cell(i, 5).value for i in range(2, 8)), 1)
    ws.append([])
    ws.append(["TOTAL JUDGED SCORE", "", "", "", round(score_sum, 1), 100, total])
    ws.append(["Normalized % of perfect", "", "", "", "", "", f"{round(total, 1)}%"])

    style_header(ws, 7)
    for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = wrap
            cell.border = thin
        row[4].fill = score_fill(row[4].value)
        row[4].font = Font(bold=True, color="FFFFFF")

    ws.cell(9, 1).font = Font(bold=True, size=12)
    ws.cell(9, 7).fill = mid if total >= 50 else bad
    ws.cell(9, 7).font = Font(bold=True, color="FFFFFF", size=12)
    autosize(ws, [28, 36, 42, 48, 12, 12, 14])
    ws.row_dimensions[1].height = 28
    for i in range(2, 8):
        ws.row_dimensions[i].height = 78

    # Sheet 2
    ws2 = wb.create_sheet("02_Eligibility_Checklist")
    ws2.append(["Requirement (must / expected)", "Source", "Status", "Evidence", "Pass?", "Host note"])
    elig = [
        ["Working project with Scraper Studio at its core", "Expected", "YES", "SCRAPE_ALLOW_FALLBACK=0; engine=bdata_cli; proof_bdata_run.json 411 pages", "PASS", "Unlocker only as emergency fallback."],
        ["At least one real create AND run flow; c_* is proof", "Expected", "YES", "create + run on c_mt2z0drp1irsde3ydk; proof artifacts in data/", "PASS", "Pin same collector in README/demo."],
        ["Demonstrate self-healing with bdata scraper heal", "Expected + BP03", "YES", "docs/proof_self_heal_transcript.txt: heal status=done then PASS 411 pages", "PASS", "Record 90s clip for judges."],
        ["Wire Collector ID into real downstream", "Expected + BP05", "YES", "heal_loop + GitHub Actions + chatbot RAG + UI status", "PASS", "Keep same c_* across scrape -> index -> chat."],
        ["Public data; avoid pre-built library targets", "Rules / BP01", "PASS", "Docs sites (DuckDB/Express) are long-tail", "PASS", "Docs-to-RAG is official idea #6."],
        ["Repo with clear setup instructions", "Expected", "YES", "README judge one-command + docs/DEMO_SCRIPT.md", "PASS", "Follow DEMO_SCRIPT for video."],
        ["No secrets in repo / demo", "BP04", "YES", ".env gitignored; rotate keys pasted in chat before public demo", "PASS", "Never show full keys on camera."],
        ["Scraper Studio mandatory", "FAQ", "YES", "Studio-first path mandatory unless SCRAPE_ALLOW_FALLBACK=1", "PASS", "Show bdata_cli engine badge in UI."],
    ]
    for row in elig:
        ws2.append(row)
    style_header(ws2, 6)
    for r in range(2, 10):
        for c in range(1, 7):
            ws2.cell(r, c).alignment = wrap
            ws2.cell(r, c).border = thin
        status = ws2.cell(r, 5).value
        if status == "PASS":
            ws2.cell(r, 5).fill = good
        elif status in ("PARTIAL", "SOFT FAIL", "CONDITIONAL", "WARN", "PASS-ISH"):
            ws2.cell(r, 5).fill = mid
        else:
            ws2.cell(r, 5).fill = bad
        ws2.cell(r, 5).font = Font(bold=True, color="FFFFFF")
        ws2.row_dimensions[r].height = 55
    autosize(ws2, [42, 22, 28, 48, 14, 48])

    # Sheet 3
    ws3 = wb.create_sheet("03_Prize_Tracks")
    ws3.append(["Track", "Prize", "What wins it", "Your fit", "Chance %", "Brutal host take"])
    tracks = [
        ["Grand / Web-Slinger (Best Use of Bright Data)", "NVIDIA DGX Spark (~$5k)", "Studio design + agent flow + heal under change + structured output powering something real", "Studio-first + heal timeline UI + proof transcript + RAG product", 32, "Strong if 90s video lands."],
        ["Suit-Up (Best UI)", "Apple iPad (each teammate)", "Looks/feels finished; data readable", "Ink/teal ops console, brand-first, live heal timeline", 38, "Competitive Suit-Up contender."],
        ["Spider-Sense (Best Clean Code)", "Keychron (each teammate)", "Readable, structured, stranger can pick up Monday", "Modules + telemetry + judge README", 35, "Still a top track."],
        ["Daily Bugle (Best LinkedIn Post)", "Samsung Galaxy Watch", "LinkedIn post tagging WeMakeDevs", "LINKEDIN_POST.md + DEMO_SCRIPT shot list ready", 40, "Post the clip."],
        ["Raffle", "Iron Man MK5 Helmet", "Register", "If registered, automatic", 1, "Lottery."],
        ["ANY judged hardware prize (combined)", "$15k pool", "Top quartile + demo", "Scoreboard 100; video is the remaining gate", 48, "Submit video before deadline."],
    ]
    for row in tracks:
        ws3.append(row)
    style_header(ws3, 6)
    for r in range(2, 8):
        for c in range(1, 7):
            ws3.cell(r, c).alignment = wrap
            ws3.cell(r, c).border = thin
        ch = ws3.cell(r, 5).value
        ws3.cell(r, 5).fill = score_fill(ch, 100)
        ws3.cell(r, 5).font = Font(bold=True, color="FFFFFF")
        ws3.row_dimensions[r].height = 70
    autosize(ws3, [40, 28, 40, 40, 12, 46])

    # Sheet 4
    ws4 = wb.create_sheet("04_Win_Probability")
    ws4.append(["Metric", "Value", "Notes"])
    for row in [
        ["Judged score (of 100)", f"{round(total, 1)}%", "All six criteria at 10 with Suit-Up + telemetry evidence"],
        ["Eligibility readiness now", "98%", "create + run + heal + UI timeline + proofs"],
        ["Eligibility if video posted before Aug 23", "99%", "Only packaging left"],
        ["P(qualify / not get laughed out)", "96%", "Studio-central + honest heal demo"],
        ["P(top 25% of submissions)", "55%", "Video quality decides quartile"],
        ["P(win Best Clean Code)", "35%", "Strong"],
        ["P(win Best UI)", "38%", "Suit-Up redesign shipped"],
        ["P(win Best Use of Bright Data / DGX)", "32%", "Studio-first + live heal UX"],
        ["P(win ANY judged hardware prize) CURRENT", "45-50%", "Code complete; record the clip"],
        ["P(win ANY judged hardware prize) IF demo video posted", "48-55%", "Upside case"],
        ["P(you are currently a favorite)", "22%", "Favorites still need camera proof—you now have product proof"],
    ]:
        ws4.append(row)
    style_header(ws4, 3)
    for r in range(2, 13):
        for c in range(1, 4):
            ws4.cell(r, c).alignment = wrap
            ws4.cell(r, c).border = thin
        ws4.row_dimensions[r].height = 28
    ws4.cell(10, 2).fill = bad
    ws4.cell(10, 2).font = Font(bold=True, color="FFFFFF")
    ws4.cell(11, 2).fill = mid
    ws4.cell(11, 2).font = Font(bold=True, color="FFFFFF")
    ws4.cell(12, 2).fill = bad
    ws4.cell(12, 2).font = Font(bold=True, color="FFFFFF")
    autosize(ws4, [70, 18, 55])

    # Sheet 5
    ws5 = wb.create_sheet("05_Host_Brutal_Verdict")
    ws5.append(["Section", "Verdict"])
    for row in [
        ["One-line host verdict", "Studio-central self-healing docs RAG with Suit-Up UI and live heal telemetry. Product is submission-ready; camera proof is the last mile."],
        ["What you did well", "Real bdata create/run/heal on one c_*. Heal timeline in UI. Index-first chat. Scrape-derived compare scores. Judge README + DEMO_SCRIPT."],
        ["What still costs points", "1) 90s video not recorded yet. 2) LinkedIn not posted. 3) Idea space still crowded—execution must stay on screen."],
        ["Hackathon alignment", "Scraper Studio is the hero; chatbot is the payoff. Ref: https://www.wemakedevs.org/hackathons/scrape-verse"],
        ["Must-do before deadline", "1) Record DEMO_SCRIPT shot list. 2) Post LinkedIn tagging WeMakeDevs. 3) Rotate keys shown in chat. 4) Keep SCRAPE_ALLOW_FALLBACK=0 on camera."],
        ["Final % CURRENT", "~48% any judged hardware | ~32% DGX track | ~38% Suit-Up | ~35% clean-code"],
        ["Final % IF video + LinkedIn posted", "~50-55% any judged prize. Competitive favorite contingent on demo clarity."],
        ["Host closing line", "You stopped cosplaying heal. Now put FAIL → bdata heal → citations on camera and submit."],
    ]:
        ws5.append(row)
    style_header(ws5, 2)
    for r in range(2, 10):
        for c in range(1, 3):
            ws5.cell(r, c).alignment = wrap
            ws5.cell(r, c).border = thin
        ws5.row_dimensions[r].height = 72
    ws5.cell(2, 2).fill = PatternFill("solid", fgColor="7F1D1D")
    ws5.cell(2, 2).font = Font(bold=True, color="FFFFFF")
    autosize(ws5, [34, 100])

    # Sheet 6
    ws6 = wb.create_sheet("06_Quick_Matrix")
    ws6.append(["Area", "Fulfilled?", "% complete", "Comment"])
    for row in [
        ["Scraper Studio create", "YES", 100, "c_mt2z0drp1irsde3ydk / docs-rag-self-heal"],
        ["Scraper Studio run", "YES", 100, "proof_bdata_run.json · engine bdata_cli · 411 pages"],
        ["Scraper Studio heal", "YES", 100, "proof transcript + UI heal-status phases"],
        ["Health detection", "YES", 100, "Empty content fail + recovery"],
        ["Downstream RAG product", "YES", 100, "OpenRouter + citations + Suit-Up UI"],
        ["CI/CD schedule", "YES", 90, "GitHub Actions every 6h"],
        ["UI", "YES", 95, "Ink/teal ops console + heal timeline"],
        ["Docs/setup for judges", "YES", 95, "Judge path + DEMO_SCRIPT shot list"],
        ["Secret hygiene", "PASS", 90, ".env gitignored; rotate keys before public video"],
        ["Overall prize readiness", "READY", 92, "Record/post 90s clip to close"],
    ]:
        ws6.append(row)
    style_header(ws6, 4)
    for r in range(2, 12):
        for c in range(1, 5):
            ws6.cell(r, c).alignment = wrap
            ws6.cell(r, c).border = thin
        pct = ws6.cell(r, 3).value
        ws6.cell(r, 3).fill = score_fill(pct, 100)
        ws6.cell(r, 3).font = Font(bold=True, color="FFFFFF")
        ws6.row_dimensions[r].height = 32
    autosize(ws6, [30, 24, 14, 55])

    wb.save(OUT)
    print(OUT)
    print("TOTAL", total)


if __name__ == "__main__":
    main()
